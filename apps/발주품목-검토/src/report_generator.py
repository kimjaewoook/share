import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
from config.settings import OUTPUT_DIR

DAY_NAMES = ['월','화','수','목','금','토','일']

COLUMN_MAP = [
    ("트리거",             "trigger_type",    "text",    "center"),
    ("요청일",             "date_requested",   "date",    "center"),
    ("팩토리",             "factory",         "text",    "center"),
    ("요청부서",           "department",       "text",    "center"),
    ("요청자",             "requestor",        "text",    "center"),
    ("입고요청일",          "expected_date",    "date",    "center"),
    ("자재코드(물류)",       "mat_code",         "text",    "center"),
    ("자재코드(구매)",       "mat_code_buy",     "text",    "center"),
    ("자재명",             "mat_name",         "text",    "left"),
    ("발주수량",           "vendor_po_qty",    "number",  "right"),
    ("박스수",             "box_count",        "decimal", "right"),
    ("박스입수량",         "box_qty",          "mixed",   "right"),
    ("파렛트수",           "pallet_count",     "decimal", "right"),
    ("PLT당 총 수량",      "plt_total_qty",    "mixed",   "right"),
    ("요청수량",           "qty",              "number",  "right"),
    ("가용재고",           "remaining_stock",  "number",  "right"),
    ("공급업체",           "vendor",           "text",    "center"),
    ("일평균(30d)",        "avg_30d",          "number",  "right"),
    ("일평균(60d)",        "avg_60d",          "number",  "right"),
    ("일평균(90d)",        "avg_90d",          "number",  "right"),
    ("일평균(8w)",         "avg_8w",           "number",  "right"),
    ("일평균(성수기)",      "avg_peak",         "number",  "right"),
    ("일평균(비수기)",      "avg_off",          "number",  "right"),
    ("채택(MAX)",          "adopted_usage",    "number",  "right"),
    ("소진 일수",          "cover_days",       "days",    "right"),
    ("소진 주수",          "cover_weeks",      "weeks",   "right"),
    ("AI 코멘트",          "action",           "text",    "left"),
]

USAGE_COL_INDICES = [18, 19, 20, 21, 22, 23]  # R~W (1-based)
PASTEL_PALETTE = [
    "FFF4E6", "E6F9FF", "E6FFE6", "FFE6E6", "F9E6FF",
    "FFFFE6", "E6FFF9", "FFE6F9", "E6E6FF", "F2FFE6",
    "E6F2FF", "FFE6EB", "E6FFED", "F4E6FF", "FFFBE6",
    "E6EFFF", "FFEDE6", "E6FFE0", "FFE0E6", "E0E6FF",
    "FDF5E6", "F0F8FF", "F5FFFA", "FFF0F5", "F5F5DC"
]
NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'
DECIMAL_FMT = '#,##0.0'
DAYS_FMT = '#,##0.0"일"'
WEEKS_FMT = '#,##0.0"주"'

HEADER_NOTES = {
    "트리거": "현장요청=현장 부서 자재요청서\n자동보충=AI 재고소진 모니터링",
    "팩토리": "자재요청서 > 팩토리 열",
    "요청부서": "자재요청서 > 팀명 열",
    "요청자": "자재요청서 > 신청인 열",
    "요청일": "자재요청서 > 요청일자 열",
    "입고요청일": "L/T + 팩토리별 입고요일 룰 기반 AI 산출",
    "자재코드(물류)": "자재요청서 > 자재코드 열",
    "자재코드(구매)": "단가테이블 > 자재코드(구매) 열\n유효여부=O 조건",
    "자재명": "자재요청서 > 자재명 열",
    "요청수량": "자재요청서 > 수량(낱개) 합산",
    "가용재고": "장현석_재고관리시트 > 잔여재고 수량",
    "발주수량": "AI 산출: (요청+L/T수요+안전재고) - 잔여재고\nMOQ/박스/파렛트 단위 올림 적용",
    "공급업체": "단가테이블 > 공급업체명(약식)\n유효여부=O + 팩토리 라우팅",
    "일평균(30d)": "자재요청서 최근 30일 요청 SUM ÷ 30",
    "일평균(60d)": "자재요청서 최근 60일 요청 SUM ÷ 60",
    "일평균(90d)": "자재요청서 최근 90일 요청 SUM ÷ 90",
    "일평균(8w)": "장현석_재고관리시트 > 1일 평균(최근 8주)",
    "일평균(성수기)": "재고소진예상시점 > 3~5,9~11월 평균\n기간: 2025.05~2026.03",
    "일평균(비수기)": "재고소진예상시점 > 6~8,12~2월 평균\n기간: 2025.05~2026.03",
    "채택(MAX)": "6개 소요량 중 최대값 자동 채택",
    "소진 일수": "발주수량 ÷ 채택(MAX)",
    "소진 주수": "소진 일수 ÷ 7",
    "AI 코멘트": "발주 판단 사유, 미납 독촉, 교차검증 경고 등",
}

def _to_date(val):
    if not val or not isinstance(val, str):
        return val
    val = str(val).strip()
    if val in ["차주 화/수", ""]:
        return val
    try:
        return datetime.strptime(val[:10], "%Y-%m-%d")
    except:
        return val

def _to_float(val):
    """값을 네이티브 Python float로 강제 변환 (numpy/pandas 타입 방지)"""
    import math
    if val is None or val == "":
        return float(0)
    try:
        n = float(str(val).replace(",",""))
        if math.isnan(n) or math.isinf(n):
            return float(0)
        return float(n)  # numpy.float64 → Python float
    except:
        return float(0)

def generate_excel_report(plan_df, po_headers=None):
    now = datetime.now()
    day_name = DAY_NAMES[now.weekday()]
    filename = f"ai발주검토_{now.strftime('%Y-%m-%d')}({day_name})_{now.strftime('%H-%M-%S')}.xlsx"
    filepath = OUTPUT_DIR / filename

    if 'vendor' in plan_df.columns:
        if 'vendor_po_qty' in plan_df.columns:
            plan_df['_needs_order'] = plan_df['vendor_po_qty'] > 0
            plan_df = plan_df.sort_values(by=['_needs_order', 'vendor', 'trigger_type', 'mat_code'], ascending=[False, True, True, True], ignore_index=True)
            plan_df = plan_df.drop(columns=['_needs_order'])
        else:
            plan_df = plan_df.sort_values(by=['vendor','trigger_type','mat_code'], ignore_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "발주제안결과"

    headers = [c[0] for c in COLUMN_MAP]
    ws.append(headers)

    # 스타일
    hdr_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    a_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    a_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    a_right  = Alignment(horizontal="right",  vertical="center")
    body_font = Font(size=10)
    body_font_gray = Font(size=10, color="888888")
    light_hdr_fill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")

    for cell in ws[1]:
        hdr_text = cell.value
        if hdr_text in ["박스입수량", "PLT당 총 수량"]:
            cell.fill = light_hdr_fill
        else:
            cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = a_center
        cell.border = border
        # 헤더 메모 삽입
        if hdr_text in HEADER_NOTES:
            cell.comment = Comment(HEADER_NOTES[hdr_text], "AI발주검토")

    # 팩토리별 전용 스타일 (배경: 은은한 파스텔톤, 글자: 채도 낮은 진한 톤 + Bold)
    FACTORY_STYLES = {
        "자재창고": {"bg": "EAEAEA", "fg": "333333"}, # 군포 키워드보다 먼저 매칭되도록 최상단 배치
        "군포": {"bg": "0054A6", "fg": "FFFFFF"},
        "성수": {"bg": "C00000", "fg": "FFFFFF"},
        "부산": {"bg": "1E8449", "fg": "FFFFFF"},
        "파주": {"bg": "E1D5E7", "fg": "660066"},
    }

    # 업체 고유 색상 매핑
    vendor_color_map = {"": "FFFFFF"}
    color_idx = 0

    for _, row in plan_df.iterrows():
        is_no_order = (row.get("vendor_po_qty", 0) <= 0)
        
        cur_vendor = row.get("vendor", "")
        if cur_vendor not in vendor_color_map:
            vendor_color_map[cur_vendor] = PASTEL_PALETTE[color_idx % len(PASTEL_PALETTE)]
            color_idx += 1
        
        if is_no_order:
            row_hex = "F2F2F2"
            row_font = Font(size=10, color="888888")
        else:
            row_hex = vendor_color_map[cur_vendor]
            row_font = body_font

        row_fill = PatternFill(start_color=row_hex,
                               end_color=row_hex, fill_type="solid")

        # 행 데이터 빌드
        current_row = ws.max_row + 1
        has_plt_qty = False
        plt_qty_val = row.get("plt_total_qty", "-")
        if plt_qty_val not in ["-", ""] and not pd.isna(plt_qty_val):
            try:
                if float(plt_qty_val) > 0:
                    has_plt_qty = True
            except:
                pass

        has_box_qty = False
        box_qty_val = row.get("box_qty", "-")
        if box_qty_val not in ["-", ""] and not pd.isna(box_qty_val):
            try:
                if float(box_qty_val) > 0:
                    has_box_qty = True
            except:
                pass

        row_data = []
        for i, (_, key, dtype, _) in enumerate(COLUMN_MAP):
            val = row.get(key, "")
            if key == "vendor_po_qty":
                if has_plt_qty:
                    val = f"=M{current_row}*N{current_row}"
                elif has_box_qty:
                    val = f"=K{current_row}*L{current_row}"
            else:
                if dtype == "date":
                    val = _to_date(val)
                elif dtype in ("number", "days", "weeks"):
                    val = _to_float(val)
                elif dtype == "mixed" and val not in ["-", ""]:
                    val = _to_float(val)
            row_data.append(val)
        ws.append(row_data)

        # 셀 스타일
        r = ws.max_row
        for col_idx, (col_name, key, dtype, align) in enumerate(COLUMN_MAP, start=1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            
            if col_name in ["박스입수량", "PLT당 총 수량"]:
                cell.font = body_font_gray
            else:
                cell.font = row_font

            if col_idx not in USAGE_COL_INDICES or is_no_order:
                cell.fill = row_fill

            # 팩토리 컬럼 디자인 강조
            if not is_no_order and key == "factory":
                fac_val = str(cell.value or "")
                for f_key, f_style in FACTORY_STYLES.items():
                    if f_key in fac_val:
                        cell.fill = PatternFill(start_color=f_style["bg"], end_color=f_style["bg"], fill_type="solid")
                        cell.font = Font(bold=True, color=f_style["fg"], size=10)
                        break

            if align == "right":
                cell.alignment = a_right
            elif align == "left":
                cell.alignment = a_left
            else:
                cell.alignment = a_center

            if dtype == "number":
                if key in ["vendor_po_qty", "qty", "remaining_stock"]:
                    cell.number_format = INT_FMT
                else:
                    cell.number_format = NUM_FMT
            elif dtype == "decimal":
                cell.number_format = DECIMAL_FMT
            elif dtype == "days":
                cell.number_format = DAYS_FMT
            elif dtype == "weeks":
                cell.number_format = WEEKS_FMT
            elif dtype == "date":
                cell.number_format = 'yyyy-mm-dd'
            elif dtype == "mixed":
                cell.number_format = INT_FMT

        # 행 기준 상대 색조: 각 행의 5개 소요량 내에서 min~max 비율로 색칠
        if not is_no_order:
            vals = []
            for ci in USAGE_COL_INDICES:
                try:
                    v = float(ws.cell(row=r, column=ci).value or 0)
                except:
                    v = 0.0
                vals.append(v)
            row_max = max(vals)
            row_min = min(vals)
            for idx, ci in enumerate(USAGE_COL_INDICES):
                cell = ws.cell(row=r, column=ci)
                v = vals[idx]
                if row_max == row_min or row_max == 0:
                    # 전부 같거나 0이면 흰색
                    ratio = 0.0
                else:
                    ratio = (v - row_min) / (row_max - row_min)
                # 0은 무조건 흰색 배경 + 연한 회색 폰트
                if v == 0:
                    hex_color = "FFFFFF"
                    cell.font = Font(size=10, color="CCCCCC")
                elif row_max == row_min:
                    # 전부 같은 값 → 연한 옐로
                    hex_color = "FFFDE7"
                else:
                    ratio = (v - row_min) / (row_max - row_min)
                    # 파스텔 그린(0) → 연한 옐로(0.5) → 파스텔 레드(1.0)
                    if ratio <= 0.5:
                        t = ratio / 0.5  # 0→1
                        red = int(210 + 45 * t)      # D2→FF  (210→255)
                        green = int(228 - 5 * t)     # E4→DF  (228→223)
                        blue = int(192 - 5 * t)      # C0→BB  (192→187)
                    else:
                        t = (ratio - 0.5) / 0.5  # 0→1
                        red = int(255)               # FF 고정
                        green = int(223 - 73 * t)    # DF→96  (223→150)
                        blue = int(187 - 47 * t)     # BB→8C  (187→140)
                    hex_color = f"{red:02X}{green:02X}{blue:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                # MAX 셀 Bold
                if v == row_max and row_max > 0:
                    cell.font = Font(bold=True, size=10)

    # 너비
    widths = {
        'A':10,'B':16,'C':13,'D':10,'E':10,'F':16,
        'G':16,'H':16,'I':38,'J':12,'K':10,'L':10,
        'M':10,'N':14,'O':12,'P':12,'Q':14,'R':11,'S':11,
        'T':11,'U':11,'V':11,'W':11,'X':12,'Y':11,'Z':11,'AA':55,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"\n✅ 엑셀 리포트 생성이 완료되었습니다.\n경로: {filepath}")
    return filepath