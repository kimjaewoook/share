#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
커머스 통합자동화 — Task 4: 데이터 가공 (task_processor.py)

source/ 폴더의 출고 파일(2개) + WMS 재고 파일(1개)을 읽어
정리 시트, WMS 양식, ERP 양식을 자동 생성하여 output/ 에 저장.

원본: py_commerce-item/automate_excel.py (587줄) 리팩토링
"""

import copy
import re
import unicodedata
import warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from config import INPUT_DIR, OUTPUT_DIR, TODAY_STR, TODAY_COMPACT, DAY_KR

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


# ──────────────────────────── 컬럼 순서 상수 ────────────────────────────

SUMMARY_COLS = [
    'concatenate', '배송시작일', '팩토리', '이관공장', '회원카드번호', '1회이용',
    '분류번호', '키커', '슬롯', '적용', '상품명', '옵션', '상품품목', 'ERP 상품명',
    '완료', '불출수량', '재고량', '로케이션', '로케이션 수', '상품코드', '상품 바코드',
    'ERP 번호', '옷걸이 바코드', '상태', '인계상태', '출고준비일시', '출고일',
    '수거신청일', '여러밤배송 여부',
]

FIRST_SHEET_COLS = [
    'concatenate', '배송시작일', '팩토리', '이관공장', '회원카드번호', '1회이용',
    '분류번호', '키커', '슬롯', '적용', '상품명', '옵션', '상품품목', '완료',
    '불출수량', '재고량', '로케이션', '상품코드', '상품 바코드', 'ERP 번호',
    '옷걸이 바코드', '상태', '인계상태', '출고준비일시', '출고일', '수거신청일',
    '여러밤배송 여부',
]

DEFAULT_WMS_COLS = [
    '회사명(필수)', '브랜드(필수)', '판매처명(필수)', '상품코드(필수)', '바코드(필수)',
    '고객사 상품코드(필수)', '상품명(필수)', '옵션1', '옵션2', '유통기한', 'LOT번호',
    '센터', 'zone', '1단계', '2단계', '가용수량(필수)', '불용수량(필수)', '수령자(필수)',
    '연락처(필수)', '주소(필수)', '출고예정일(필수)', '전달방식(필수)', '비고',
]

DEFAULT_ERP_HEADER_ROWS = [
    ['품번', '조정일자', '창고코드', '장소코드', '조정구분', '조정수량', '담당자코드', '관리구분코드', '비고(건)', '비고(내역)'],
    ['ITEM_CD', 'ADJUST_DT', 'WH_CD', 'LC_CD', 'ADJUST_FG', 'ADJUST_QT', 'PLN_CD', 'MGMT_CD', 'REMARK_DC', 'REMARKD_DC'],
    [
        '타입 : 문자\n길이 : 30\n필수 : True\n설명 : 영문/숫자 기준 30자리(최대)를 입력 하세요.',
        '타입 : 문자\n길이 : 8\n필수 : True\n설명 : 영문/숫자 기준 8자리(YYYYMMDD)를 입력 하세요.',
        '타입 : 문자\n길이 : 4\n필수 : True\n설명 : 영문/숫자 기준 4자리(최대)를 입력 하세요.',
        '타입 : 문자\n길이 : 4\n필수 : True\n설명 : 영문/숫자 기준 4자리(최대)를 입력 하세요.',
        '타입 : 문자\n길이 : 1\n필수 : True\n설명 : 0.기초조정,1.입고조정,2.출고조정',
        '타입 : 숫자\n길이 : 17,6\n필수 : True\n설명 : 숫자 기준 17,6자리를 입력 하세요.(숫자만 입력(소숫점자리수 정확히 입력))',
        '타입 : 문자\n길이 : 5\n필수 : False\n설명 : 영문/숫자 기준 5자리(최대)를 입력 하세요.',
        '타입 : 문자\n길이 : 10\n필수 : False\n설명 : 성수 20000, 군포 30000',
        '타입 : 문자\n길이 : 60\n필수 : False\n설명 : 영문/숫자 기준 60자리(최대)를 입력 하세요.',
        '타입 : 문자\n길이 : 60\n필수 : False\n설명 : 영문/숫자 기준 60자리(최대)를 입력 하세요.',
    ],
]


# ──────────────────────────── 유틸리티 ────────────────────────────

def _ensure_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """누락 컬럼 NaN 추가 후 순서대로 정렬."""
    for c in cols:
        if c not in df.columns:
            df[c] = np.nan
    return df[cols]


def _calc_actual_out(inv_qty, req_qty):
    """재고량 vs 불출수량 → 실제 출고 가능 수량."""
    try:
        return inv_qty if float(inv_qty) < float(req_qty) else req_qty
    except (ValueError, TypeError):
        return req_qty


def _extract_target_date(file_paths: list[Path]) -> datetime:
    """파일명에서 YYYY-MM-DD 날짜 추출 → 최신 날짜 반환."""
    all_dates = []
    for f in file_paths:
        matches = re.findall(r'(\d{4}-\d{2}-\d{2})', f.name)
        all_dates.extend(matches)

    if all_dates:
        return datetime.strptime(sorted(all_dates)[-1], "%Y-%m-%d")
    return datetime.now()


# ──────────────────────────── 1단계: 파일 읽기 ────────────────────────────

def _read_source_files() -> tuple[pd.DataFrame, pd.DataFrame, list[Path]] | None:
    """source/ 폴더에서 출고 파일(product_release_*) + WMS 재고 파일(wms_inventory_*) 읽기.

    Returns
    -------
    (df_first, df_inv, source_files) 또는 None (파일 부족 시)
    """
    release_files: list[Path] = []
    inventory_files: list[Path] = []

    for f in INPUT_DIR.iterdir():
        if f.suffix != '.xlsx' or f.name.startswith('~') or f.name.startswith('.'):
            continue
        norm_name = unicodedata.normalize('NFC', f.name)

        if norm_name.startswith("product_release_"):
            release_files.append(f)
        elif norm_name.startswith("wms_inventory_") or norm_name.startswith("상품별 재고"):
            inventory_files.append(f)

    if len(release_files) < 1:
        print(f"  [오류] 출고 파일이 없습니다. (찾은 수: {len(release_files)})")
        return None
    if len(inventory_files) < 1:
        print(f"  [오류] WMS 재고 파일이 없습니다. (찾은 수: {len(inventory_files)})")
        return None

    inventory_file = sorted(inventory_files, key=lambda p: p.name)[-1]

    print(f"    📄 출고 파일: {[f.name for f in release_files]}")
    print(f"    📄 재고 파일: {inventory_file.name}")

    # 출고 파일 병합
    df_first = pd.concat(
        [pd.read_excel(f, engine='openpyxl', dtype={'상품 바코드': str}) for f in release_files],
        ignore_index=True,
    )

    # '공장' → '팩토리' 변환
    if '공장' in df_first.columns:
        df_first.rename(columns={'공장': '팩토리'}, inplace=True)
        df_first['팩토리'] = df_first['팩토리'].replace({'군포공장': '군포팩토리', '성수공장': '성수팩토리'})
    elif '팩토리' not in df_first.columns:
        df_first['팩토리'] = ''

    df_first['상품 바코드'] = df_first['상품 바코드'].fillna('').astype(str)
    df_first['concatenate'] = df_first['팩토리'].fillna('') + df_first['상품 바코드']

    # '위치' → '완료'
    if '위치' in df_first.columns:
        df_first.rename(columns={'위치': '완료'}, inplace=True)

    # WMS 재고 파일 읽기 — 시트명 자동 탐색
    df_inv = None
    try:
        df_inv = pd.read_excel(
            inventory_file, sheet_name='상품 재고', engine='openpyxl',
            dtype={'바코드': str, 'zone': str, '1단계': str, '2단계': str},
        )
    except ValueError:
        # '상품 재고' 시트가 없으면 첫 번째 시트로 폴백
        print("    ℹ️ '상품 재고' 시트 미발견 → 첫 번째 시트 사용")
        df_inv = pd.read_excel(
            inventory_file, sheet_name=0, engine='openpyxl',
            dtype={'바코드': str, 'zone': str, '1단계': str, '2단계': str},
        )

    if '바코드' in df_inv.columns:
        df_inv['바코드'] = df_inv['바코드'].fillna('').astype(str)
    else:
        print("    ⚠️ 재고 파일에 '바코드' 컬럼이 없습니다.")

    all_files = release_files + [inventory_file]
    return df_first, df_inv, all_files


# ──────────────────────────── 2단계: 정리 시트 생성 ────────────────────────────

def _build_summary(df_first: pd.DataFrame, df_inv: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """출고 + 재고 매칭 → '정리' 시트용 DataFrame."""
    print("    🔄 정리 시트 데이터 계산 중...")

    df_summary = df_first.drop_duplicates(subset=['concatenate'], keep='first').copy()

    if '완료' not in df_summary.columns:
        df_summary['완료'] = ''
    df_summary['완료'] = df_summary['완료'].astype(object)

    has_barcode = '바코드' in df_inv.columns

    # 바코드별 가용재고 합계
    inv_sum_dict = {}
    if has_barcode and '가용 재고' in df_inv.columns:
        inv_sum_dict = df_inv.groupby('바코드')['가용 재고'].sum().to_dict()

    # 바코드별 로케이션 수
    inv_count_dict = {}
    if has_barcode:
        inv_count_dict = df_inv['바코드'].value_counts().to_dict()

    # 바코드별 첫 번째 로케이션/상품코드/상품명
    inv_info_dict = {}
    if has_barcode:
        info_cols = [c for c in ['zone', '1단계', '2단계', '상품코드', '상품명'] if c in df_inv.columns]
        if info_cols:
            df_first_info = df_inv[df_inv['바코드'] != ''].groupby('바코드')[info_cols].first()
            for bcode, row in df_first_info.iterrows():
                z = str(row.get('zone', '')) if pd.notna(row.get('zone')) else ''
                s1 = str(row.get('1단계', '')) if pd.notna(row.get('1단계')) else ''
                s2 = str(row.get('2단계', '')) if pd.notna(row.get('2단계')) else ''
                pcode = str(row.get('상품코드', '')) if pd.notna(row.get('상품코드')) else ''
                pname = str(row.get('상품명', '')) if pd.notna(row.get('상품명')) else ''
                inv_info_dict[bcode] = {'loc': z + s1 + s2, 'code': pcode, 'name': pname}

    # 출고 카운트
    release_counts = df_first['concatenate'].value_counts().to_dict()
    barcode_release_counts = df_first['상품 바코드'].value_counts().to_dict()

    # 벡터화 매핑
    df_summary['불출수량'] = df_summary['concatenate'].map(release_counts).fillna(0).astype(int)
    df_summary['재고량'] = df_summary['상품 바코드'].map(inv_sum_dict).fillna(0)
    df_summary['로케이션 수'] = df_summary['상품 바코드'].map(inv_count_dict).fillna(0).astype(int)

    df_summary['로케이션'] = df_summary['상품 바코드'].map(
        lambda b: inv_info_dict.get(b, {}).get('loc', '')
    )
    df_summary['상품코드'] = df_summary['상품 바코드'].map(
        lambda b: inv_info_dict.get(b, {}).get('code', '')
    )
    df_summary['ERP 상품명'] = df_summary['상품 바코드'].map(
        lambda b: inv_info_dict.get(b, {}).get('name', '')
    )

    # 총 재고 부족 여부 (벡터화)
    total_release = df_summary['상품 바코드'].map(barcode_release_counts).fillna(0)
    total_inv = df_summary['상품 바코드'].map(inv_sum_dict).fillna(0)
    is_short = total_release > total_inv

    existing_vals = df_summary['완료'].fillna('').astype(str).str.strip()
    short_mask = is_short & ~existing_vals.str.contains('⚠️총 재고 부족', na=False)

    df_summary.loc[short_mask & (existing_vals == ''), '완료'] = '⚠️총 재고 부족'
    df_summary.loc[short_mask & (existing_vals != ''), '완료'] = (
        existing_vals[short_mask & (existing_vals != '')] + ' ⚠️총 재고 부족'
    )

    # 컬럼 순서 + 정렬
    df_summary = _ensure_columns(df_summary, SUMMARY_COLS)
    df_first = _ensure_columns(df_first, FIRST_SHEET_COLS)
    df_summary = df_summary.sort_values(by=['팩토리', '로케이션'], ascending=[True, True], na_position='last')

    return df_summary, df_first


# ──────────────────────────── 3단계: 엑셀 저장 + 서식 ────────────────────────────

def _apply_formatting(output_path: Path) -> None:
    """정리 시트에 팩토리별 색상, 폰트, 경고 표시 적용."""
    try:
        wb = load_workbook(output_path)
        if '정리' not in wb.sheetnames:
            return

        ws = wb['정리']

        # ── 서식 참조 파일 탐색 (source/ 내 '커머스' 키워드 파일) ──
        ref_fill, ref_font = None, None
        for f in INPUT_DIR.iterdir():
            norm_name = unicodedata.normalize('NFC', f.name)
            if "커머스" in norm_name and f.suffix == ".xlsx":
                try:
                    ref_wb = load_workbook(f, data_only=True)
                    if '정리' in ref_wb.sheetnames:
                        ref_cell = ref_wb['정리'].cell(row=1, column=1)
                        ref_fill = copy.copy(ref_cell.fill)
                        ref_font = copy.copy(ref_cell.font)
                except Exception:
                    pass
                break

        # 헤더 서식 적용
        if ref_fill or ref_font:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                if ref_fill:
                    cell.fill = ref_fill
                if ref_font:
                    cell.font = ref_font

        # 스타일 정의
        font_blue_18_bold = Font(size=18, bold=True, color="0000FF")
        font_red_14_bold = Font(size=14, bold=True, color="FF0000")
        font_black_14_bold = Font(size=14, bold=True, color="000000")

        fill_pastel_green = PatternFill(start_color="E2EFDA", fill_type="solid")
        fill_pastel_blue = PatternFill(start_color="DDEBF7", fill_type="solid")
        fill_dark_green = PatternFill(start_color="C6E0B4", fill_type="solid")
        fill_dark_blue = PatternFill(start_color="BDD7EE", fill_type="solid")

        # 열 인덱스 매핑
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                col_map[val] = col

        ci_qty = col_map.get('불출수량')
        ci_inv = col_map.get('재고량')
        ci_barcode = col_map.get('상품 바코드')
        ci_loc = col_map.get('로케이션')
        ci_factory = col_map.get('팩토리')
        ci_complete = col_map.get('완료')

        max_col = ws.max_column
        for row_idx in range(2, ws.max_row + 1):
            is_short = False
            if ci_qty and ci_inv:
                v_qty = ws.cell(row=row_idx, column=ci_qty).value
                v_inv = ws.cell(row=row_idx, column=ci_inv).value
                if pd.notna(v_qty) and pd.notna(v_inv):
                    try:
                        is_short = float(v_inv) < float(v_qty)
                    except ValueError:
                        pass

            if ci_factory:
                factory_val = ws.cell(row=row_idx, column=ci_factory).value
                fill = None
                if factory_val == "성수팩토리":
                    fill = fill_dark_green if is_short else fill_pastel_green
                elif factory_val == "군포팩토리":
                    fill = fill_dark_blue if is_short else fill_pastel_blue
                if fill:
                    for c in range(1, max_col + 1):
                        ws.cell(row=row_idx, column=c).fill = fill

            if ci_qty:
                ws.cell(row=row_idx, column=ci_qty).font = font_blue_18_bold
            if ci_barcode:
                ws.cell(row=row_idx, column=ci_barcode).font = font_red_14_bold
            if ci_loc:
                ws.cell(row=row_idx, column=ci_loc).font = font_black_14_bold

            if ci_complete:
                val = ws.cell(row=row_idx, column=ci_complete).value
                if pd.notna(val) and "⚠️" in str(val):
                    ws.cell(row=row_idx, column=ci_complete).font = font_red_14_bold

        wb.save(output_path)

    except Exception as e:
        print(f"    ⚠️ 서식 적용 중 오류: {e}")


# ──────────────────────────── 4단계: WMS 양식 생성 ────────────────────────────

def _generate_wms(df_summary: pd.DataFrame, target_date_str: str) -> None:
    """WMS 간편출고등록 양식 생성 (DEFAULT 상수 사용)."""
    print("    📝 WMS 업로드 양식 생성 중...")

    wms_cols = DEFAULT_WMS_COLS
    default_wms = {
        '회사명(필수)': '의식주컴퍼니', '브랜드(필수)': '의식주컴퍼니',
        '수령자(필수)': '의식주컴퍼니', '연락처(필수)': '010-9999-9999',
        '주소(필수)': '의식주컴퍼니', '전달방식(필수)': '기타', '비고': '군포/성수 커머스',
    }

    try:
        df_valid = df_summary[pd.to_numeric(df_summary['재고량'], errors='coerce').fillna(0) > 0].copy()

        wms_data = []
        for _, row in df_valid.iterrows():
            new_row = {c: default_wms.get(c, '') for c in wms_cols}
            new_row['회사명(필수)'] = '의식주컴퍼니'
            new_row['브랜드(필수)'] = '의식주컴퍼니'
            new_row['판매처명(필수)'] = row.get('팩토리', '') or '의식주컴퍼니'
            new_row['상품코드(필수)'] = row.get('상품코드', '')
            new_row['바코드(필수)'] = row.get('상품 바코드', '')
            new_row['고객사 상품코드(필수)'] = row.get('ERP 번호', '')
            new_row['상품명(필수)'] = row.get('ERP 상품명', '')
            new_row['가용수량(필수)'] = _calc_actual_out(row.get('재고량', 0), row.get('불출수량', 0))
            new_row['불용수량(필수)'] = 0
            new_row['출고예정일(필수)'] = target_date_str
            wms_data.append(new_row)

        df_wms = pd.DataFrame(wms_data, columns=wms_cols)
        wms_path = OUTPUT_DIR / f"{TODAY_COMPACT}({DAY_KR}) 간편출고등록양식(WMS).xlsx"
        df_wms.to_excel(wms_path, index=False)
        print(f"    📥 WMS 양식 → {wms_path.name}")

    except Exception as e:
        print(f"    ⚠️ WMS 양식 생성 실패: {e}")


# ──────────────────────────── 5단계: ERP 양식 생성 ────────────────────────────

def _generate_erp(df_summary: pd.DataFrame) -> None:
    """더존 ERP 업로드 양식 생성 (DEFAULT 상수 사용)."""
    print("    📝 더존 ERP 업로드 양식 생성 중...")

    erp_header_rows = DEFAULT_ERP_HEADER_ROWS
    erp_cols = erp_header_rows[0]

    col_idx_map = {}
    for i, name in enumerate(erp_cols):
        if name not in col_idx_map:
            col_idx_map[name] = i

    try:
        df_valid = df_summary[pd.to_numeric(df_summary['재고량'], errors='coerce').fillna(0) > 0].copy()

        erp_data = []
        for _, row in df_valid.iterrows():
            new_row = [''] * len(erp_cols)

            if '품번' in col_idx_map:
                new_row[col_idx_map['품번']] = row.get('ERP 번호', '')
            if '조정일자' in col_idx_map:
                new_row[col_idx_map['조정일자']] = TODAY_COMPACT
            if '창고코드' in col_idx_map:
                new_row[col_idx_map['창고코드']] = '1000'
            if '장소코드' in col_idx_map:
                new_row[col_idx_map['장소코드']] = '1000'
            if '조정구분' in col_idx_map:
                new_row[col_idx_map['조정구분']] = '2'
            if '조정수량' in col_idx_map:
                new_row[col_idx_map['조정수량']] = _calc_actual_out(row.get('재고량', 0), row.get('불출수량', 0))
            if '담당자코드' in col_idx_map:
                new_row[col_idx_map['담당자코드']] = '11111'
            if '관리구분코드' in col_idx_map:
                fac = row.get('팩토리', '')
                new_row[col_idx_map['관리구분코드']] = (
                    '20000' if fac == '성수팩토리' else ('30000' if fac == '군포팩토리' else '')
                )
            if '비고(건)' in col_idx_map:
                new_row[col_idx_map['비고(건)']] = '군포/성수 커머스'
            if '비고(내역)' in col_idx_map:
                new_row[col_idx_map['비고(내역)']] = row.get('회원카드번호', '')

            erp_data.append(new_row)

        final_data = erp_header_rows + erp_data
        df_erp = pd.DataFrame(final_data)
        erp_path = OUTPUT_DIR / f"{TODAY_COMPACT}({DAY_KR}) 간편출고등록양식(더존ERP).xlsx"
        df_erp.to_excel(erp_path, index=False, header=False, engine='openpyxl')
        print(f"    📥 ERP 양식 → {erp_path.name}")

    except Exception as e:
        print(f"    ⚠️ ERP 양식 생성 실패: {e}")


# ──────────────────────────── 공개 API ────────────────────────────

def run_processing() -> tuple[bool, "Path | None"]:
    """Task 4 실행: source/ 파일 읽기 → 정리 → WMS/ERP 양식 → output/ 저장.

    Returns
    -------
    (성공 여부, 최종 엑셀 파일 경로 또는 None)
    """
    print("\n" + "=" * 60)
    print("📋 Task 4: 데이터 가공 (Pandas)")
    print("=" * 60)

    # 1. 파일 읽기
    result = _read_source_files()
    if result is None:
        return False, None

    df_first, df_inv, source_files = result

    # 2. 정리 시트 생성
    df_summary, df_first = _build_summary(df_first, df_inv)

    # 3. 날짜 정보
    target_date_obj = _extract_target_date(source_files)
    target_date_str = target_date_obj.strftime("%Y-%m-%d")

    # 4. 메인 엑셀 저장 (정리 + 첫번째 시트 + 상품 재고)
    output_filename = f"{TODAY_COMPACT}({DAY_KR}) 커머스 출고.xlsx"
    output_path = OUTPUT_DIR / output_filename

    print(f"    💾 '{output_filename}' 저장 중...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='정리', index=False)
        df_first.to_excel(writer, sheet_name='첫번째 시트', index=False)
        df_inv.to_excel(writer, sheet_name='상품 재고', index=False)

    # 5. 서식 적용
    _apply_formatting(output_path)
    print(f"  ✅ 최종 결과물 → {output_path.name}")

    # 6. WMS / ERP 양식 생성
    _generate_wms(df_summary, target_date_str)
    _generate_erp(df_summary)

    return True, output_path
