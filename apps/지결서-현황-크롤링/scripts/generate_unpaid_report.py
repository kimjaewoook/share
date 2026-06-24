import sys
import os
import re
import datetime
import colorsys
import pandas as pd
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import gspread
from config import SERVICE_ACCOUNT_FILE
from collections import defaultdict
from openpyxl.styles import PatternFill, Font

def parse_amount(val):
    if not val: return 0
    try:
        val_str = str(val).replace(',', '').split('.')[0]
        return int(val_str)
    except:
        return 0

def is_unshaded(cell):
    fmt = cell.get("effectiveFormat", {}).get("backgroundColorStyle", {}).get("rgbColor", {})
    if not fmt:
        return True
    if fmt.get("red", 0) == 1 and fmt.get("green", 0) == 1 and fmt.get("blue", 0) == 1:
        return True
    return False

def extract_date_from_subject(subject):
    if not isinstance(subject, str): return ""
    m = re.search(r'\((1?[0-9]/[1-3]?[0-9])\)', subject)
    if m:
        try:
            month, day = map(int, m.group(1).split('/'))
            return f"{datetime.datetime.now().year}-{month:02d}-{day:02d}"
        except:
            return ""
    return ""

def run():
    gc = gspread.service_account(filename=str(SERVICE_ACCOUNT_FILE))
    SHEET_ID = "1OW405qapjFE5pPWaXKfqHPQ72KdHg33eO47Q7VpQAec"
    url = f"https://sheets.googleapis.com/v4/spreadsheets/{SHEET_ID}?includeGridData=true&ranges=A:Z"
    
    print("구글 시트 데이터 가져오는 중...")
    resp = gc.http_client.request('get', url).json()
    
    row_data = resp['sheets'][0]['data'][0].get('rowData', [])
    if not row_data:
        return
        
    headers = [c.get('formattedValue', '') for c in row_data[0].get('values', [])]
    vendor_idx = headers.index('vendor_name')
    amount_idx = headers.index('amount')
    subject_idx = headers.index('subject')
    doc_url_idx = headers.index('doc_url')
    detail_idx = headers.index('detail_comments')

    unpaid_data = defaultdict(list)
    
    for row in row_data[1:]:
        cells = row.get('values', [])
        if not cells: continue
        if not is_unshaded(cells[0]): continue
            
        def get_val(idx):
            if idx < len(cells): return cells[idx].get('formattedValue', '').strip()
            return ""
            
        vendor = get_val(vendor_idx)
        if not vendor: continue
            
        subj_val = get_val(subject_idx)
        amount_val = parse_amount(get_val(amount_idx))
        detail_val = get_val(detail_idx)
        
        orig_amount = amount_val
        unpaid_amount = amount_val
        extra_note = ""
        
        if vendor == "메이네티":
            if "1월 비닐" in subj_val and "18,435,460" in detail_val:
                unpaid_amount = 18435460
                extra_note = "2026.04.30 지급완료 57,436,270원 (미지급잔액 18,435,460원)"
            elif "12월 비닐" in subj_val:
                extra_note = detail_val.replace('\n', ' / ')
            
        item = {
            "date": extract_date_from_subject(subj_val),
            "subject": subj_val,
            "orig_amount": orig_amount,
            "unpaid_amount": unpaid_amount,
            "doc_url": get_val(doc_url_idx),
            "extra_note": extra_note
        }
        unpaid_data[vendor].append(item)
        
    # 상계 처리 로직
    for vendor in unpaid_data:
        items = unpaid_data[vendor]
        items.sort(key=lambda x: x['date'])
        
        sched_payments = []
        if vendor == "메이네티":
            sched_payments = [
                {'col': '6/19예정', 'amount': 100_000_000},
                {'col': '7/17예정', 'amount': 100_000_000}
            ]
        
        for item in items:
            unpaid = item['unpaid_amount']
            item['6/19예정'] = ""
            item['7/17예정'] = ""
            
            if vendor == "메이네티":
                for p in sched_payments:
                    item[p['col']] = 0
                    if p['amount'] > 0 and unpaid > 0:
                        apply = min(p['amount'], unpaid)
                        item[p['col']] = apply
                        p['amount'] -= apply
                        unpaid -= apply
                        
            item['잔여 미수'] = unpaid

    output_rows = []
    sorted_vendors = sorted(unpaid_data.keys())
    
    for vendor in sorted_vendors:
        items = unpaid_data[vendor]
        
        for item in items:
            output_rows.append({
                '업체명': vendor,
                '지급요청일': item['date'],
                '내역 (제목)': item['subject'],
                '계산서 금액': item['orig_amount'],
                '미결제 금액': item['unpaid_amount'],
                '6/19예정': item['6/19예정'],
                '7/17예정': item['7/17예정'],
                '잔여 미수': item['잔여 미수'],
                '기안문서 URL': item['doc_url'],
                '비고': item['extra_note']
            })

    df = pd.DataFrame(output_rows)
    out_path = Path(__file__).resolve().parent.parent / "output" / "unpaid_report.xlsx"
    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='미납현황')
        worksheet = writer.sheets['미납현황']
        
        widths = {'A':15, 'B':15, 'C':60, 'D':16, 'E':16, 'F':16, 'G':16, 'H':16, 'I':40, 'J':40}
        for col, w in widths.items():
            worksheet.column_dimensions[col].width = w
            
        unique_vendors = list(df['업체명'].dropna().unique())
        n_vendors = len(unique_vendors)
        
        vendor_color_map = {}
        for i, v in enumerate(unique_vendors):
            # 인접한 업체끼리 색상이 비슷해지는 그라데이션 현상을 막기 위해 황금비(Golden Ratio) 활용
            hue = (i * 0.618033988749895) % 1.0
            # 밝기 92%, 채도 70%의 부드러운 파스텔톤
            rgb = colorsys.hls_to_rgb(hue, 0.92, 0.70)
            vendor_color_map[v] = f"{int(rgb[0]*255):02X}{int(rgb[1]*255):02X}{int(rgb[2]*255):02X}"
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            vendor_val = row[0].value
            if not vendor_val: continue
            
            base_vendor = str(vendor_val).strip()
            color_hex = vendor_color_map.get(base_vendor, "FFFFFF")
                
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            
            for cell in row:
                cell.fill = fill
                
                # 금액 열 포맷 (D~H)
                if cell.column_letter in ['D', 'E', 'F', 'G', 'H']:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                        
                # 하이퍼링크 (I열)
                if cell.column_letter == 'I' and cell.value and "http" in str(cell.value):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

    print(f"엑셀 리포트가 성공적으로 생성되었습니다: {out_path}")

if __name__ == "__main__":
    run()
